"""
Audit Log Module for Standings Adjustments
Maintains a single ongoing Excel file that tracks all adjustment history.
"""

import os
import json
import pandas as pd
import time
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

CACHE_DIR = Path(__file__).resolve().parent.parent.parent / "data" / "weekly_standings_cache"
AUDIT_DIR = Path(__file__).resolve().parent.parent.parent / "data" / "audit_logs"


def ensure_audit_dir():
	"""Creates the audit directory if it doesn't exist."""
	AUDIT_DIR.mkdir(parents=True, exist_ok=True)


def get_audit_log_path(league_id, league_name=None):
	"""Returns the path to the ongoing audit log for a league."""
	ensure_audit_dir()
	league_label = league_name.replace(' ', '_') if league_name else league_id
	filename = f"adjustment_audit_log_{league_label}.xlsx"
	return AUDIT_DIR / filename


def load_cached_period_data(league_id, period):
	"""Loads cached data for a specific period."""
	cache_file = CACHE_DIR / f"weekly_standings_{league_id}_{period}.json"
	
	if not cache_file.exists():
		return None, None
	
	try:
		with open(cache_file, 'r') as f:
			cache_data = json.load(f)
			df = pd.DataFrame.from_records(cache_data['data'])
			min_games = cache_data.get('metadata', {}).get('min_games', 35)
			return df, min_games
	except (json.JSONDecodeError, KeyError):
		return None, None


def calculate_adjustment_details(raw_data, min_games):
	"""Calculates detailed adjustment information for audit logging."""
	df = pd.DataFrame.from_records(raw_data) if isinstance(raw_data, list) else raw_data.copy()
	
	# Calculate FP/G
	df['Calc_FPG'] = df.apply(
		lambda row: round(row['FPts'] / row['GP'], 2) if row['GP'] > 0 else 0,
		axis=1
	)
	
	# Calculate games over limit
	df['Games_Over'] = df['GP'].apply(lambda x: max(0, x - min_games))
	
	# Calculate adjustment
	df['Adjustment_Amount'] = (df['Games_Over'] * df['Calc_FPG']).round(2)
	
	# Calculate adjusted values
	df['Adjusted_FPts'] = (df['FPts'] - df['Adjustment_Amount']).round(2)
	
	# Calculate final submission value (negative, rounded)
	df['Submitted_Value'] = -(df['Adjustment_Amount'].round(0).astype(int))
	
	# Calculate percentage impact
	df['Adjustment_Pct'] = (
		(df['Adjustment_Amount'] / df['FPts'] * 100)
		.round(2)
		.fillna(0)
	)
	
	return df


def create_new_audit_log(filepath, league_id, league_name):
	"""Creates a new audit log Excel file with proper structure."""
	workbook = openpyxl.Workbook()
	
	# Remove default sheet
	if 'Sheet' in workbook.sheetnames:
		workbook.remove(workbook['Sheet'])
	
	# Create Adjustment History sheet
	history_sheet = workbook.create_sheet("Adjustment History", 0)
	
	# Title
	history_sheet['A1'] = f"Standings Adjustment Audit Log - {league_name or league_id}"
	history_sheet['A1'].font = Font(size=16, bold=True)
	history_sheet.merge_cells('A1:M1')
	
	# Metadata
	history_sheet['A3'] = "Created:"
	history_sheet['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
	history_sheet['A4'] = "League ID:"
	history_sheet['B4'] = league_id
	
	# Headers
	headers = [
		'Timestamp', 'Period', 'Team Name', 'Team ID',
		'Min Games', 'GP', 'Original FPts', 'Calc FP/G',
		'Games Over', 'Adjustment Amount', 'Adjustment %',
		'Adjusted FPts', 'Submitted Value'
	]
	
	for col, header in enumerate(headers, start=1):
		cell = history_sheet.cell(row=6, column=col, value=header)
		cell.font = Font(bold=True, color="FFFFFF")
		cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
		cell.alignment = Alignment(horizontal="center", wrap_text=True)
	
	# Set column widths
	column_widths = {
		'A': 20, 'B': 8, 'C': 30, 'D': 20,
		'E': 10, 'F': 8, 'G': 12, 'H': 10,
		'I': 10, 'J': 15, 'K': 12, 'L': 12, 'M': 15
	}
	for col, width in column_widths.items():
		history_sheet.column_dimensions[col].width = width
	
	# Create Summary sheet
	summary_sheet = workbook.create_sheet("Summary", 1)
	
	summary_sheet['A1'] = "Adjustment Summary by Period"
	summary_sheet['A1'].font = Font(size=14, bold=True)
	summary_sheet.merge_cells('A1:G1')
	
	summary_headers = [
		'Period', 'Date Processed', 'Min Games', 'Teams Adjusted',
		'Total Adjustment', 'Avg Adjustment', 'Max Adjustment'
	]
	
	for col, header in enumerate(summary_headers, start=1):
		cell = summary_sheet.cell(row=3, column=col, value=header)
		cell.font = Font(bold=True, color="FFFFFF")
		cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
		cell.alignment = Alignment(horizontal="center")
	
	for col in range(1, 8):
		summary_sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
	
	workbook.save(filepath)
	return workbook


def append_adjustment_to_log(league_id, period, league_name=None, notes=None):
	"""
	Appends a new adjustment entry to the ongoing audit log.
	
	Args:
		league_id: Fantrax league ID
		period: Scoring period number
		league_name: Optional friendly league name
		notes: Optional notes about this adjustment
	
	Returns:
		Tuple of (success: bool, message: str, filepath: Path)
	"""
	try:
		# Load cached data
		raw_df, min_games = load_cached_period_data(league_id, period)
		
		if raw_df is None:
			return False, f"No cached data found for period {period}", None
		
		# Calculate adjustments
		adj_df = calculate_adjustment_details(raw_df, min_games)
		
		# Get audit log path
		filepath = get_audit_log_path(league_id, league_name)
		
		# Create new file if it doesn't exist
		if not filepath.exists():
			create_new_audit_log(filepath, league_id, league_name)
		
		# Open workbook
		workbook = openpyxl.load_workbook(filepath)
		history_sheet = workbook["Adjustment History"]
		summary_sheet = workbook["Summary"]
		
		# Find next row in history sheet
		next_row = history_sheet.max_row + 1
		
		# Add timestamp
		timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
		
		# Append each team's adjustment
		for _, team in adj_df.iterrows():
			history_sheet.cell(row=next_row, column=1, value=timestamp)
			history_sheet.cell(row=next_row, column=2, value=period)
			history_sheet.cell(row=next_row, column=3, value=team['team_name'])
			history_sheet.cell(row=next_row, column=4, value=team.get('team_id', 'N/A'))
			history_sheet.cell(row=next_row, column=5, value=min_games)
			history_sheet.cell(row=next_row, column=6, value=int(team['GP']))
			history_sheet.cell(row=next_row, column=7, value=float(team['FPts']))
			history_sheet.cell(row=next_row, column=8, value=float(team['Calc_FPG']))
			history_sheet.cell(row=next_row, column=9, value=int(team['Games_Over']))
			
			# Highlight adjustment if > 0
			adj_cell = history_sheet.cell(row=next_row, column=10, value=float(team['Adjustment_Amount']))
			if team['Adjustment_Amount'] > 0:
				adj_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
				adj_cell.font = Font(bold=True)
			
			history_sheet.cell(row=next_row, column=11, value=f"{float(team['Adjustment_Pct'])}%")
			history_sheet.cell(row=next_row, column=12, value=float(team['Adjusted_FPts']))
			history_sheet.cell(row=next_row, column=13, value=int(team['Submitted_Value']))
			
			next_row += 1
		
		# Update summary sheet
		summary_row = summary_sheet.max_row + 1
		teams_adjusted = (adj_df['Adjustment_Amount'] > 0).sum()
		total_adj = adj_df['Adjustment_Amount'].sum()
		avg_adj = adj_df['Adjustment_Amount'].mean()
		max_adj = adj_df['Adjustment_Amount'].max()
		
		summary_sheet.cell(row=summary_row, column=1, value=period)
		summary_sheet.cell(row=summary_row, column=2, value=timestamp)
		summary_sheet.cell(row=summary_row, column=3, value=min_games)
		summary_sheet.cell(row=summary_row, column=4, value=teams_adjusted)
		summary_sheet.cell(row=summary_row, column=5, value=round(total_adj, 2))
		summary_sheet.cell(row=summary_row, column=6, value=round(avg_adj, 2))
		summary_sheet.cell(row=summary_row, column=7, value=round(max_adj, 2))
		
		# Save workbook
		workbook.save(filepath)
		
		return True, f"Successfully logged adjustments for Period {period}", filepath
		
	except Exception as e:
		return False, f"Error appending to audit log: {e}", None


def get_audit_log_info(league_id, league_name=None):
	"""
	Returns information about the audit log for a league.
	
	Returns:
		Dict with log info or None if doesn't exist
	"""
	filepath = get_audit_log_path(league_id, league_name)
	
	if not filepath.exists():
		return None
	
	try:
		workbook = openpyxl.load_workbook(filepath, read_only=True)
		summary_sheet = workbook["Summary"]
		
		# Count periods logged
		periods_logged = summary_sheet.max_row - 3  # Subtract header rows
		
		result = {
			'filepath': filepath,
			'exists': True,
			'periods_logged': max(0, periods_logged),
			'file_size': filepath.stat().st_size
		}
		
		# Explicitly close the workbook
		workbook.close()
		
		return result
	except Exception:
		return None


def reset_audit_log(league_id, league_name=None):
	"""
	Deletes the audit log file for a league.
	WARNING: This permanently deletes all audit history!
	
	Args:
		league_id: Fantrax league ID
		league_name: Optional friendly league name
	
	Returns:
		Tuple of (success: bool, message: str)
	"""
	try:
		filepath = get_audit_log_path(league_id, league_name)
		
		if not filepath.exists():
			return False, "No audit log exists to reset."
		
		# Give Windows a moment to release any file locks
		time.sleep(0.5)
		
		# Try to delete the file with retry logic
		max_attempts = 3
		for attempt in range(max_attempts):
			try:
				filepath.unlink()
				return True, f"Audit log successfully deleted: {filepath.name}"
			except PermissionError:
				if attempt < max_attempts - 1:
					time.sleep(1)  # Wait longer between retries
				else:
					raise
		
	except Exception as e:
		return False, f"Error deleting audit log: {e}"

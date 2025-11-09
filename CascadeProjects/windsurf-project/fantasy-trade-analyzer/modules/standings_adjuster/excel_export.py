"""
Excel Export Module for Standings Tools
Generates comprehensive Excel workbooks with historical tracking of all adjustments.
"""

import os
import json
import pandas as pd
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

CACHE_DIR = Path(__file__).resolve().parent.parent.parent / "data" / "weekly_standings_cache"
EXPORT_DIR = Path(__file__).resolve().parent.parent.parent / "data" / "standings_exports"


def ensure_export_dir():
	"""Creates the export directory if it doesn't exist."""
	EXPORT_DIR.mkdir(parents=True, exist_ok=True)


def load_all_cached_data(league_id):
	"""
	Loads all cached standings data for a given league ID.
	Returns a dictionary mapping period -> data.
	"""
	if not CACHE_DIR.exists():
		return {}
	
	cached_data = {}
	for file_path in CACHE_DIR.glob(f"weekly_standings_{league_id}_*.json"):
		try:
			period = int(file_path.stem.split('_')[-1])
			with open(file_path, 'r') as f:
				data = json.load(f)
				cached_data[period] = data
		except (ValueError, json.JSONDecodeError, KeyError):
			continue
	
	return cached_data


def calculate_all_adjustments(raw_data, min_games):
	"""
	Calculates all adjustment-related values for a given dataset.
	Returns a DataFrame with comprehensive adjustment tracking.
	"""
	df = pd.DataFrame.from_records(raw_data)
	
	# Original values
	df['Original FPts'] = df['FPts']
	df['Original GP'] = df['GP']
	df['Original Rank'] = df['rank']
	
	# Calculate FP/G
	df['Calc FP/G'] = df.apply(
		lambda row: row['FPts'] / row['GP'] if row['GP'] > 0 else 0, 
		axis=1
	)
	
	# Calculate games over limit
	df['Min Games Limit'] = min_games
	df['Games Over'] = df['GP'].apply(lambda x: max(0, x - min_games))
	
	# Calculate adjustment
	df['Adjustment Amount'] = (df['Games Over'] * df['Calc FP/G']).round(2)
	
	# Calculate adjusted values
	df['Adjusted FPts'] = (df['FPts'] - df['Adjustment Amount']).round(2)
	
	# Calculate final submission value (negative, rounded)
	df['Final Adjustment (Submitted)'] = -(df['Adjustment Amount'].round(0).astype(int))
	
	# Calculate adjusted rank
	df_sorted = df.sort_values('Adjusted FPts', ascending=False).reset_index(drop=True)
	df_sorted['Adjusted Rank'] = range(1, len(df_sorted) + 1)
	df = df.merge(df_sorted[['team_name', 'Adjusted Rank']], on='team_name', how='left')
	
	# Calculate rank change
	df['Rank Change'] = df['Original Rank'].astype(int) - df['Adjusted Rank'].astype(int)
	
	# Percentage impact
	df['Adjustment %'] = (
		(df['Adjustment Amount'] / df['Original FPts'] * 100)
		.round(2)
		.fillna(0)
	)
	
	return df


def create_summary_sheet(workbook, all_periods_data, league_id):
	"""Creates a summary sheet with overview of all periods."""
	ws = workbook.create_sheet("Summary", 0)
	
	# Title
	ws['A1'] = f"Standings Adjustments Summary - League {league_id}"
	ws['A1'].font = Font(size=16, bold=True)
	ws.merge_cells('A1:G1')
	
	# Export metadata
	ws['A3'] = "Export Date:"
	ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
	ws['A4'] = "Total Periods:"
	ws['B4'] = len(all_periods_data)
	
	# Summary table headers
	headers = [
		"Period", "Min Games", "Teams", "Total Adjustments", 
		"Avg Adjustment", "Max Adjustment", "Teams Affected"
	]
	
	for col, header in enumerate(headers, start=1):
		cell = ws.cell(row=6, column=col, value=header)
		cell.font = Font(bold=True)
		cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
		cell.font = Font(color="FFFFFF", bold=True)
		cell.alignment = Alignment(horizontal="center")
	
	# Populate summary data
	row = 7
	for period in sorted(all_periods_data.keys()):
		data = all_periods_data[period]
		min_games = data['metadata'].get('min_games', 'N/A')
		df = pd.DataFrame.from_records(data['data'])
		
		# Calculate stats
		df['adjustment_calc'] = df.apply(
			lambda r: max(0, r['GP'] - min_games) * (r['FPts'] / r['GP']) if r['GP'] > 0 else 0,
			axis=1
		)
		
		total_adj = df['adjustment_calc'].sum()
		avg_adj = df['adjustment_calc'].mean()
		max_adj = df['adjustment_calc'].max()
		teams_affected = (df['adjustment_calc'] > 0).sum()
		
		ws.cell(row=row, column=1, value=period)
		ws.cell(row=row, column=2, value=min_games)
		ws.cell(row=row, column=3, value=len(df))
		ws.cell(row=row, column=4, value=round(total_adj, 2))
		ws.cell(row=row, column=5, value=round(avg_adj, 2))
		ws.cell(row=row, column=6, value=round(max_adj, 2))
		ws.cell(row=row, column=7, value=teams_affected)
		
		row += 1
	
	# Auto-adjust column widths
	for col in range(1, 8):
		ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18


def create_period_detail_sheet(workbook, period, data):
	"""Creates a detailed sheet for a specific period."""
	min_games = data['metadata'].get('min_games', 35)
	df = calculate_all_adjustments(data['data'], min_games)
	
	# Sort by adjusted rank
	df = df.sort_values('Adjusted Rank')
	
	ws = workbook.create_sheet(f"Period {period}")
	
	# Title
	ws['A1'] = f"Period {period} - Detailed Adjustments"
	ws['A1'].font = Font(size=14, bold=True)
	ws.merge_cells('A1:K1')
	
	# Metadata
	ws['A3'] = "Minimum Games Limit:"
	ws['B3'] = min_games
	ws['B3'].font = Font(bold=True)
	
	# Column headers
	columns = [
		'Team Name', 'Original Rank', 'Adjusted Rank', 'Rank Change',
		'Original FPts', 'GP', 'Calc FP/G', 'Games Over',
		'Adjustment Amount', 'Adjustment %', 'Adjusted FPts', 
		'Final Adjustment (Submitted)'
	]
	
	for col, header in enumerate(columns, start=1):
		cell = ws.cell(row=5, column=col, value=header)
		cell.font = Font(bold=True)
		cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
		cell.font = Font(color="FFFFFF", bold=True)
		cell.alignment = Alignment(horizontal="center", wrap_text=True)
	
	# Data rows
	row = 6
	for _, team_row in df.iterrows():
		ws.cell(row=row, column=1, value=team_row['team_name'])
		ws.cell(row=row, column=2, value=int(team_row['Original Rank']))
		ws.cell(row=row, column=3, value=int(team_row['Adjusted Rank']))
		
		# Rank change with color coding
		rank_change = int(team_row['Rank Change'])
		cell = ws.cell(row=row, column=4, value=rank_change)
		if rank_change > 0:
			cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
			cell.font = Font(color="006100")
		elif rank_change < 0:
			cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
			cell.font = Font(color="9C0006")
		
		ws.cell(row=row, column=5, value=round(team_row['Original FPts'], 2))
		ws.cell(row=row, column=6, value=int(team_row['Original GP']))
		ws.cell(row=row, column=7, value=round(team_row['Calc FP/G'], 2))
		ws.cell(row=row, column=8, value=int(team_row['Games Over']))
		
		# Adjustment amount with highlighting
		adj_cell = ws.cell(row=row, column=9, value=round(team_row['Adjustment Amount'], 2))
		if team_row['Adjustment Amount'] > 0:
			adj_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
			adj_cell.font = Font(color="9C5700", bold=True)
		
		ws.cell(row=row, column=10, value=f"{round(team_row['Adjustment %'], 2)}%")
		ws.cell(row=row, column=11, value=round(team_row['Adjusted FPts'], 2))
		ws.cell(row=row, column=12, value=int(team_row['Final Adjustment (Submitted)']))
		
		row += 1
	
	# Auto-adjust column widths
	for col in range(1, 13):
		ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
	ws.column_dimensions['A'].width = 30  # Team name column wider


def create_comparison_sheet(workbook, all_periods_data):
	"""Creates a sheet comparing adjustments across all periods."""
	ws = workbook.create_sheet("Cross-Period Comparison")
	
	# Title
	ws['A1'] = "Cross-Period Team Comparison"
	ws['A1'].font = Font(size=14, bold=True)
	ws.merge_cells('A1:F1')
	
	# Collect all unique teams
	all_teams = set()
	for data in all_periods_data.values():
		df = pd.DataFrame.from_records(data['data'])
		all_teams.update(df['team_name'].tolist())
	
	all_teams = sorted(all_teams)
	
	# Create comparison table
	periods = sorted(all_periods_data.keys())
	
	# Headers
	ws['A3'] = "Team Name"
	ws['A3'].font = Font(bold=True)
	
	col = 2
	for period in periods:
		ws.cell(row=3, column=col, value=f"P{period} Adj")
		ws.cell(row=3, column=col).font = Font(bold=True)
		ws.cell(row=3, column=col).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
		ws.cell(row=3, column=col).font = Font(color="FFFFFF", bold=True)
		col += 1
	
	ws.cell(row=3, column=col, value="Total Adj")
	ws.cell(row=3, column=col).font = Font(bold=True, color="FFFFFF")
	ws.cell(row=3, column=col).fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
	
	# Data rows
	row = 4
	for team in all_teams:
		ws.cell(row=row, column=1, value=team)
		
		total_adj = 0
		col = 2
		for period in periods:
			data = all_periods_data[period]
			min_games = data['metadata'].get('min_games', 35)
			df = calculate_all_adjustments(data['data'], min_games)
			
			team_data = df[df['team_name'] == team]
			if not team_data.empty:
				adj = team_data['Adjustment Amount'].iloc[0]
				ws.cell(row=row, column=col, value=round(adj, 2))
				total_adj += adj
			else:
				ws.cell(row=row, column=col, value=0)
			
			col += 1
		
		# Total column
		total_cell = ws.cell(row=row, column=col, value=round(total_adj, 2))
		if total_adj > 0:
			total_cell.font = Font(bold=True)
		
		row += 1
	
	# Auto-adjust column widths
	ws.column_dimensions['A'].width = 30
	for col in range(2, len(periods) + 3):
		ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12


def create_raw_data_sheet(workbook, all_periods_data):
	"""Creates a sheet with all raw scraped data."""
	ws = workbook.create_sheet("Raw Data Archive")
	
	ws['A1'] = "Raw Scraped Data - All Periods"
	ws['A1'].font = Font(size=14, bold=True)
	ws.merge_cells('A1:H1')
	
	# Combine all data
	all_rows = []
	for period in sorted(all_periods_data.keys()):
		data = all_periods_data[period]
		df = pd.DataFrame.from_records(data['data'])
		df['Period'] = period
		df['Min Games'] = data['metadata'].get('min_games', 'N/A')
		all_rows.append(df)
	
	if all_rows:
		combined_df = pd.concat(all_rows, ignore_index=True)
		
		# Reorder columns
		priority_cols = ['Period', 'Min Games', 'rank', 'team_name', 'team_id', 'FPts', 'GP', 'FP/G']
		other_cols = [col for col in combined_df.columns if col not in priority_cols]
		combined_df = combined_df[priority_cols + other_cols]
		
		# Write to sheet
		for r_idx, row in enumerate(dataframe_to_rows(combined_df, index=False, header=True), start=3):
			for c_idx, value in enumerate(row, start=1):
				cell = ws.cell(row=r_idx, column=c_idx, value=value)
				if r_idx == 3:  # Header row
					cell.font = Font(bold=True)
					cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
					cell.font = Font(color="FFFFFF", bold=True)
		
		# Auto-adjust column widths
		for col in range(1, len(combined_df.columns) + 1):
			ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15


def generate_comprehensive_excel(league_id, league_name=None):
	"""
	Generates a comprehensive Excel workbook with all historical adjustment data.
	
	Args:
		league_id: The Fantrax league ID
		league_name: Optional friendly name for the league
	
	Returns:
		Path to the generated Excel file
	"""
	ensure_export_dir()
	
	# Load all cached data
	all_periods_data = load_all_cached_data(league_id)
	
	if not all_periods_data:
		raise ValueError(f"No cached data found for league {league_id}")
	
	# Create workbook
	workbook = openpyxl.Workbook()
	workbook.remove(workbook.active)  # Remove default sheet
	
	# Create all sheets
	create_summary_sheet(workbook, all_periods_data, league_id)
	
	for period in sorted(all_periods_data.keys()):
		create_period_detail_sheet(workbook, period, all_periods_data[period])
	
	create_comparison_sheet(workbook, all_periods_data)
	create_raw_data_sheet(workbook, all_periods_data)
	
	# Save workbook
	timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
	league_label = league_name.replace(' ', '_') if league_name else league_id
	filename = f"standings_adjustments_{league_label}_{timestamp}.xlsx"
	filepath = EXPORT_DIR / filename
	
	workbook.save(filepath)
	
	return filepath


def generate_period_excel(league_id, period, league_name=None):
	"""
	Generates an Excel workbook for a single period.
	
	Args:
		league_id: The Fantrax league ID
		period: The scoring period
		league_name: Optional friendly name for the league
	
	Returns:
		Path to the generated Excel file
	"""
	ensure_export_dir()
	
	# Load cached data for the period
	cache_file = CACHE_DIR / f"weekly_standings_{league_id}_{period}.json"
	
	if not cache_file.exists():
		raise ValueError(f"No cached data found for league {league_id}, period {period}")
	
	with open(cache_file, 'r') as f:
		data = json.load(f)
	
	# Create workbook
	workbook = openpyxl.Workbook()
	workbook.remove(workbook.active)
	
	# Create period detail sheet
	create_period_detail_sheet(workbook, period, data)
	
	# Save workbook
	timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
	league_label = league_name.replace(' ', '_') if league_name else league_id
	filename = f"standings_adjustments_{league_label}_P{period}_{timestamp}.xlsx"
	filepath = EXPORT_DIR / filename
	
	workbook.save(filepath)
	
	return filepath
